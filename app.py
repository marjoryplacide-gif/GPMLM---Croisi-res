# =============================================================================
# APPLICATION STREAMLIT — Planning des escales paquebots — GPMLM
# =============================================================================
# Ce fichier contient TOUT le code de l'application : la logique de
# traitement des données ET l'interface utilisateur.
#
# Pour lancer l'application en local :
#     streamlit run app.py
#
# Pour mettre à jour l'application en ligne : modifier ce fichier puis le
# republier sur GitHub. Streamlit Cloud le recharge automatiquement.
# =============================================================================


# -----------------------------------------------------------------------------
# IMPORTS — les outils dont on a besoin
# -----------------------------------------------------------------------------

import datetime as dt   # pour manipuler les dates
import io               # pour créer le fichier Excel en mémoire (sans l'écrire sur le disque)

import pandas as pd     # pour manipuler les tableaux de données (comme Excel mais en code)
import streamlit as st  # pour construire l'interface web

# openpyxl : la librairie qui sait créer de vrais fichiers .xlsx
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


# -----------------------------------------------------------------------------
# CONFIGURATION DE LA PAGE
# -----------------------------------------------------------------------------

st.set_page_config(
    page_title="Planning escales paquebots - GPMLM",
    layout="centered",
)


# =============================================================================
# PARTIE 1 — CONSTANTES ET ERREUR PERSONNALISÉE
# =============================================================================

# Table de correspondance entre les noms de postes tels qu'ils apparaissent
# dans le CSV VIGIEsip, et les codes courts utilisés dans le planning final.
# Si VIGIEsip change un jour le nom d'un poste, c'est ici qu'il faut
# le mettre à jour.
POSTE_VIGIE_VERS_CODE = {
    "POINTE SIMON EST":  "CROE",
    "POINTE SIMON OUEST": "CROW",
    "TOURELLES":         "TOUR",
    "QUAI DES ANNEXES":  "ANNX",
}

# Note : le fichier final d'Yvon mentionne aussi les postes GQ, SP et BEACH,
# mais ces postes n'apparaissent pas dans l'extraction VIGIEsip.
# Leur source est à confirmer avec Yvon/Lauriane.



# Erreur personnalisée pour les problèmes liés au fichier déposé.
# Quand le fichier n'a pas le bon format, on "lève" cette erreur avec un
# message clair. Ce message s'affiche ensuite en rouge dans l'appli.
# C'est différent d'un bug dans le code : ici, c'est l'utilisateur
# qui a déposé le mauvais fichier, pas un problème technique.
# Si tu veux modifier le texte affiché, cherche : raise ErreurFichier("...")
class ErreurFichier(Exception):
    pass


# =============================================================================
# PARTIE 2 — LECTURE ET NETTOYAGE DU CSV VIGIESIP
# =============================================================================

def lire_csv_vigie(fichier):
    """
    Lit le fichier CSV exporté depuis VIGIEsip.

    Deux particularités de ce fichier à gérer :
    - Le séparateur est un point-virgule (;) et non une virgule
    - L'encodage est "cp1252" (format Windows), pas le format standard UTF-8

    Si le fichier déposé ne correspond pas à ces caractéristiques,
    ou s'il manque des colonnes attendues, on affiche un message clair
    à l'utilisateur plutôt qu'un message d'erreur technique incompréhensible.
    """
    # Tentative de lecture du CSV avec les bons paramètres
    try:
        df = pd.read_csv(fichier, sep=";", encoding="cp1252")
    except Exception:
        raise ErreurFichier(
            "Ce fichier n'a pas pu être lu. Vérifiez qu'il s'agit bien "
            "du fichier .csv exporté depuis VIGIEsip "
            "(Agenda > Prévisions Paquebots > Édition CSV)."
        )

    # Nettoyage des noms de colonnes : suppression des espaces parasites
    # (certains exports CSV ajoutent des espaces invisibles autour des noms)
    df.columns = [c.strip() for c in df.columns]

    # Vérification que toutes les colonnes attendues sont bien présentes
    # Si une colonne manque, ce n'est probablement pas le bon fichier
    colonnes_attendues = {
        "N°", "Sens", "Navire", "Armateur", "Agent",
        "Arrivée Rade", "Poste", "Dép quai", "Validé Agent", "Tête de ligne"
    }
    manquantes = colonnes_attendues - set(df.columns)
    if manquantes:
        raise ErreurFichier(
            "Ce fichier n'a pas le bon format. Vérifiez qu'il s'agit bien "
            "du fichier .csv exporté depuis VIGIEsip "
            "(Agenda > Prévisions Paquebots > Édition CSV), et pas d'un autre fichier."
        )

    return df


# =============================================================================
# PARTIE 3 — RECONSTRUCTION DES ESCALES
# =============================================================================

def construire_escales(df):
    """
    Dans le CSV VIGIEsip, chaque escale est représentée par DEUX lignes :
    - une ligne "DPQ" (Départ Pour Quai) qui contient l'heure d'arrivée
    - une ligne "DS"  (Départ du port) qui contient l'heure de départ

    Cette fonction recolle ces deux lignes en UNE SEULE par escale,
    ce qui correspond au travail manuel décrit dans le mode opératoire d'Yvon
    ("décaler la colonne Dép quai d'une ligne vers le haut").

    On travaille par paires consécutives (ligne 0+1, ligne 2+3, ligne 4+5...)
    plutôt que par le N° du navire, car un même navire peut faire plusieurs
    escales dans la saison et avoir le même N°.
    """

    df = df.reset_index(drop=True)

    # Vérification : le nombre de lignes doit être pair (toujours des paires)
    if len(df) % 2 != 0:
        raise ErreurFichier(
            "Le fichier semble incomplet ou abîmé (nombre de lignes inhabituel). "
            "Réessayez avec une nouvelle extraction depuis VIGIEsip."
        )

    # Séparation en deux tableaux : les lignes paires (arrivées) et impaires (départs)
    arrivees = df.iloc[0::2].reset_index(drop=True)  # lignes 0, 2, 4, 6...
    departs  = df.iloc[1::2].reset_index(drop=True)  # lignes 1, 3, 5, 7...

    # Vérification que chaque paire correspond bien au même navire
    sens_ok   = (arrivees["Sens"] == "DPQ").all() and (departs["Sens"] == "DS").all()
    navire_ok = (arrivees["Navire"].values == departs["Navire"].values).all()
    if not (sens_ok and navire_ok):
        raise ErreurFichier(
            "Le contenu du fichier ne correspond pas à ce qui était attendu. "
            "Réessayez avec une nouvelle extraction depuis VIGIEsip, "
            "ou contactez Marjo si le problème persiste."
        )

    # Construction du tableau final : une ligne par escale
    # On pioche les colonnes dans "arrivees" ou "departs" selon le cas
    escales = pd.DataFrame({
        "Navire":      arrivees["Navire"],
        "Armateur":    arrivees["Armateur"],
        "Agent":       arrivees["Agent"],
        "TeteLigne":   arrivees["Tête de ligne"],   # True/False dans le CSV
        "Loa":         arrivees["L"],               # Longueur du navire
        "Lar":         arrivees["l"],               # Largeur du navire
        "Te":          arrivees["TE"],              # Tirant d'eau
        "Pax":         arrivees["Pax"],             # Nombre de passagers
        "Crew":        arrivees["Crew"],            # Nombre de membres d'équipage
        "Poste":       arrivees["Poste"],           # Nom du quai (sera recodé ensuite)
        "ValideAgent": arrivees["Validé Agent"],    # Oui / Non / Inconnu
        "Obs":          arrivees["Obs"],            # Observations 
        # pd.to_datetime convertit le texte "12/10/2026 09:00" en vraie date Python
        "ArriveeRade": pd.to_datetime(
            arrivees["Arrivée Rade"], format="%d/%m/%Y %H:%M", errors="coerce"
        ),
        "DepQuai": pd.to_datetime(
            departs["Dép quai"], format="%d/%m/%Y %H:%M", errors="coerce"
        ),
    })

    return escales


# =============================================================================
# PARTIE 4 — FILTRAGE ET RECODAGE
# =============================================================================

def filtrer_escales_validees(escales):
    """
    Supprime les escales refusées par l'agent (Validé Agent = "Non").
    Les escales "Oui" et "Inconnu" sont conservées.
    C'est l'équivalent du filtre décrit dans le mode opératoire d'Yvon :
    "Supprimer toutes les lignes contenant NON".
    """
    return escales[escales["ValideAgent"] != "Non"].reset_index(drop=True)


def recoder_postes(escales):
    """
    Remplace les noms de postes complets (ex: "POINTE SIMON EST") par
    leurs codes courts (ex: "CROE"), comme demandé dans le mode opératoire.
    Si un poste du fichier n'est pas dans la table de correspondance,
    on le conserve tel quel et on avertit l'utilisateur.
    Retourne le tableau mis à jour ET la liste des postes non reconnus.
    """
    escales = escales.copy()

    # .map() applique la table de correspondance à toute la colonne d'un coup
    escales["PosteCode"] = escales["Poste"].map(POSTE_VIGIE_VERS_CODE)

    # Postes qui n'ont pas trouvé de correspondance (valeur vide après .map())
    inconnus = sorted(
        escales[escales["PosteCode"].isna()]["Poste"].dropna().unique().tolist()
    )
    if inconnus:
        # On conserve le nom d'origine pour ne pas perdre l'information
        escales.loc[escales["PosteCode"].isna(), "PosteCode"] = escales["Poste"]

    return escales, inconnus


# =============================================================================
# PARTIE 5 — DÉTECTION AUTOMATIQUE DE LA PÉRIODE
# =============================================================================

def detecter_periode(escales):
    """
    Détecte automatiquement les dates de début et fin de saison à partir
    des dates d'arrivée présentes dans le fichier.

    Une saison croisières va toujours du 1er octobre au 30 septembre suivant.
    On regarde la date la plus ancienne du fichier pour savoir
    à quelle saison elle appartient.

    Exemple : si la première date est le 12 octobre 2026,
    la saison va du 01/10/2026 au 30/09/2027.
    """
    premiere_date = escales["ArriveeRade"].min().date()

    # Si on est en octobre ou après → la saison a commencé cette année
    # Si on est entre janvier et septembre → la saison a commencé l'année d'avant
    annee_debut = premiere_date.year if premiere_date.month >= 10 else premiere_date.year - 1

    date_debut = dt.date(annee_debut, 10, 1)
    date_fin   = dt.date(annee_debut + 1, 9, 30)

    return date_debut, date_fin


# =============================================================================
# PARTIE 6 — CONSTRUCTION DU TABLEAU FINAL
# =============================================================================

def construire_tableau(escales, date_debut, date_fin):
    """
    Construit le tableau final destiné au fichier Excel :
    - Une ligne par escale réelle (pas de lignes vides)
    - Trié par date d'arrivée
    - Filtré sur la période demandée
    - Avec une colonne "Jour" (Lundi, Mardi...) pour faciliter le filtrage
    - Avec la colonne "Tête de ligne" convertie de True/False en OUI/NON
    """
    escales = escales.copy()

    # On extrait juste la date (sans l'heure) pour pouvoir comparer avec date_debut/date_fin
    escales["JourArrivee"] = escales["ArriveeRade"].dt.date

    # Filtre sur la période : on ne garde que les escales dans la plage demandée
    escales = escales[
        (escales["JourArrivee"] >= date_debut) &
        (escales["JourArrivee"] <= date_fin)
    ]

    # Tri par date d'arrivée, du plus ancien au plus récent
    escales = escales.sort_values("ArriveeRade").reset_index(drop=True)

    # Construction ligne par ligne du tableau final
    lignes = []
    for _, esc in escales.iterrows():

        

        # Conversion de True/False (CSV) en OUI/NON (lisible)
        tete_ligne = "OUI" if esc["TeteLigne"] else "NON"

        lignes.append({
            "Date":          esc["JourArrivee"],
            "Poste":         esc["PosteCode"],
            "Navire":        esc["Navire"],
            "Compagnie":     esc["Armateur"],
            "ETA":           esc["ArriveeRade"],
            "ETD":           esc["DepQuai"],
            "Agent":         esc["Agent"],
            "Tete_ligne":    tete_ligne,       # nom sans accent ni espace pour pandas
            "Loa":           esc["Loa"],
            "Lar":           esc["Lar"],
            "Te":            esc["Te"],
            "Pax":           esc["Pax"],
            "Crew":          esc["Crew"],
            "Obs":           esc["Obs"],
        })

    return pd.DataFrame(lignes)


# =============================================================================
# PARTIE 7 — GÉNÉRATION DU FICHIER EXCEL
# =============================================================================

def formater_heure(valeur):
    """Convertit une date complète en texte "08h00". Retourne None si vide."""
    if pd.isna(valeur):
        return None
    return valeur.strftime("%Hh%M")


def generer_excel(tableau, titre, date_maj):
    """
    Crée le fichier Excel final mis en forme et le retourne en mémoire
    (sous forme de bytes, prêt à être téléchargé depuis Streamlit).

    Le fichier contient :
    - Un titre et une date de mise à jour en haut
    - Un tableau structuré Excel (avec les filtres/tri automatiques)
    - Une ligne par escale réelle, triée par date
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Escales"

    # --- Styles ---
    style_titre     = Font(name="Calibri", size=14, bold=True)
    style_sous_titre = Font(name="Calibri", size=10, italic=True, color="666666")
    style_entete    = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    style_normal    = Font(name="Calibri", size=10)
    fond_entete     = PatternFill("solid", fgColor="1F4E78")  # bleu foncé
    centrer         = Alignment(horizontal="center")

    # --- Titre et date de mise à jour ---
    ws["A1"] = titre
    ws["A1"].font = style_titre

    date_premiere = tableau["Date"].min()
    date_derniere = tableau["Date"].max()

    ws["A2"] = (
        f"Du {date_premiere.strftime('%d/%m/%Y')} au {date_derniere.strftime('%d/%m/%Y')}"
        f" — Généré le {date_maj.strftime('%d/%m/%Y')} — {len(tableau)} escales"
    )
    ws["A2"].font = style_sous_titre

    # --- En-têtes des colonnes ---
    # Ce sont les libellés visibles dans Excel (peuvent contenir espaces/accents)
    entetes = [
        "Date", "Poste", "Navire", "Compagnie",
        "ETA", "ETD", "Agent", "Tête de ligne",
        "Loa (m)", "Lar (m)", "Te (m)", "Pax", "Crew", "Observations"
    ]
    ligne_entete = 4  # les en-têtes sont sur la ligne 4 (lignes 1-3 = titre + vide)

    for col_idx, libelle in enumerate(entetes, start=1):
        c = ws.cell(row=ligne_entete, column=col_idx, value=libelle)
        c.font      = style_entete
        c.fill      = fond_entete
        c.alignment = centrer

    # --- Données ---
    for i, row in enumerate(tableau.itertuples(index=False), start=ligne_entete + 1):
        eta_val = formater_heure(row.ETA) if not isinstance(row.ETA, str) else row.ETA
        etd_val = formater_heure(row.ETD) if not isinstance(row.ETD, str) else row.ETD

        valeurs = [
            row.Date, row.Poste, row.Navire, row.Compagnie,
            eta_val, etd_val, row.Agent, row.Tete_ligne,
            row.Loa, row.Lar, row.Te, row.Pax, row.Crew, row.Obs,
        ]
        for col_idx, val in enumerate(valeurs, start=1):
            c = ws.cell(row=i, column=col_idx, value=val)
            c.font = style_normal
            # Colonne Date : format jj/mm/aaaa
            if col_idx == 1 and val is not None:
                c.number_format = '[$-fr-FR]dddd d mmmm yyyy'

    # --- Tableau structuré Excel (les petites flèches de filtre/tri) ---
    derniere_ligne = ligne_entete + len(tableau)
    derniere_col   = get_column_letter(len(entetes))
    plage = f"A{ligne_entete}:{derniere_col}{derniere_ligne}"

    table = Table(displayName="EscalesPaquebots", ref=plage)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showRowStripes=True,       # alternance de couleurs pour mieux lire
        showFirstColumn=False,
        showLastColumn=False,
        showColumnStripes=False,
    )
    ws.add_table(table)

    # --- Largeurs des colonnes ---
    largeurs = [26, 9, 25, 25, 8, 8, 10, 12, 9, 9, 8, 8, 8, 30]
    for col_idx, largeur in enumerate(largeurs, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = largeur

    # Figer la première ligne de données pour garder les en-têtes visibles en scrollant
    ws.freeze_panes = f"A{ligne_entete + 1}"

    # --- Export en mémoire (pas sur le disque) ---
    # io.BytesIO() crée un "faux fichier" en mémoire.
    # Streamlit peut ensuite proposer ce contenu au téléchargement directement.
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# =============================================================================
# PARTIE 8 — FONCTION PRINCIPALE (chef d'orchestre)
# =============================================================================

def generer_planning(fichier, titre=None, date_debut=None, date_fin=None):
    """
    Fonction principale appelée par l'interface quand on clique sur "Générer".
    Elle appelle toutes les étapes dans l'ordre :
    1. Lire le CSV
    2. Reconstruire les escales (fusionner les paires arrivée/départ)
    3. Filtrer les escales non validées
    4. Recoder les postes (POINTE SIMON EST → CROE, etc.)
    5. Détecter la période de saison (sauf si précisée manuellement)
    6. Construire le tableau final
    7. Générer le fichier Excel

    Retourne : (contenu Excel en bytes, nombre d'escales, postes inconnus,
                date de début, date de fin)
    """
    df        = lire_csv_vigie(fichier)
    escales   = construire_escales(df)
    escales   = filtrer_escales_validees(escales)
    escales, postes_inconnus = recoder_postes(escales)

    # Détection automatique de la période, sauf si l'utilisateur a défini des dates manuelles
    date_debut_auto, date_fin_auto = detecter_periode(escales)
    if date_debut is None:
        date_debut = date_debut_auto
    if date_fin is None:
        date_fin = date_fin_auto

    # Titre automatique si non précisé
    if titre is None:
        titre = f"SAISON CROISIERES {date_debut.year}-{date_fin.year} - ESCALES GPMLM"

    tableau       = construire_tableau(escales, date_debut, date_fin)
    contenu_excel = generer_excel(tableau, titre, date_maj=dt.date.today())

    return contenu_excel, len(tableau), postes_inconnus, date_debut, date_fin


# =============================================================================
# PARTIE 9 — INTERFACE STREAMLIT
# =============================================================================
# Tout ce qui suit construit l'interface visible dans le navigateur.
# st.title(), st.subheader(), st.file_uploader()... sont des composants
# Streamlit qui s'affichent automatiquement dans l'ordre où ils apparaissent.

st.title("Planning des escales paquebots")
st.caption(
    "Génère automatiquement le fichier Excel destiné aux agents, "
    "à partir d'une extraction VIGIEsip (Prévisions Paquebots)."
)

st.divider()

# --- Zone de dépôt du fichier ---
st.subheader("1. Déposez votre extraction VIGIEsip")

fichier = st.file_uploader(
    "Fichier .csv exporté depuis VIGIEsip "
    "(Agenda > Prévisions Paquebots > Édition CSV)",
    type=["csv"],
)

titre_saisi = st.text_input(
    "Titre du document (laisser vide pour un titre automatique)",
    value="",
)

# --- Options avancées (repliées par défaut, pour ne pas surcharger l'interface) ---
with st.expander("Afficher seulement une période précise (optionnel)"):
    st.caption(
        "Par défaut, le planning couvre toute la saison détectée "
        "automatiquement depuis le fichier. Cochez cette case si vous "
        "voulez un planning limité à une période plus courte "
        "(ex : juste un mois, ou les prochaines semaines)."
    )
    dates_manuelles = st.checkbox("Limiter à une période précise")
    date_debut_manuelle = None
    date_fin_manuelle   = None
    if dates_manuelles:
        col1, col2 = st.columns(2)
        with col1:
            date_debut_manuelle = st.date_input("Date de début", format="DD/MM/YYYY")
        with col2:
            date_fin_manuelle = st.date_input("Date de fin", format="DD/MM/YYYY")

st.divider()

# --- Bouton de génération (visible seulement si un fichier est déposé) ---
if fichier is None:
    st.info("Déposez un fichier CSV ci-dessus pour activer la génération.")
else:
    if st.button("Générer le planning", type="primary", use_container_width=True):
        try:
            with st.spinner("Génération du fichier en cours..."):
                titre = titre_saisi.strip() or None
                contenu, nb_escales, postes_inconnus, date_debut, date_fin = generer_planning(
                    fichier,
                    titre=titre,
                    date_debut=date_debut_manuelle,
                    date_fin=date_fin_manuelle,
                )

            st.success(
                f"Fichier généré — {nb_escales} escales, "
                f"du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}."
            )

            # Avertissement si des postes du fichier n'ont pas été reconnus
            if postes_inconnus:
                st.warning(
                    "Poste(s) non reconnu(s), conservé(s) tel(s) quel(s) dans le fichier : "
                    f"{', '.join(postes_inconnus)}. "
                    "Si ce n'est pas normal, vérifier la table POSTE_VIGIE_VERS_CODE "
                    "en haut du fichier app.py."
                )

            nom_fichier = f"Planning_escales_{date_debut.year}-{date_fin.year}.xlsx"
            st.download_button(
                label="Télécharger le fichier Excel",
                data=contenu,
                file_name=nom_fichier,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                on_click=st.rerun,
            )
        # Erreur liée au fichier déposé (mauvais format, fichier abîmé...)
        except ErreurFichier as e:
            st.error(str(e))

        # Erreur technique inattendue (bug dans le code)
        except Exception as e:
            st.error(
                f"Une erreur inattendue s'est produite : {e} — "
                "Contactez Marjo si le problème persiste."
            )

st.divider()

# --- Aide repliée en bas de page ---
with st.expander("Comment obtenir le fichier CSV depuis VIGIEsip ?"):
    st.markdown("""
    1. Sur VIGIEsip, cliquer sur **Agenda** puis **Prévisions Paquebots**
    2. Définir les dates de l'extraction (une saison va du 1er octobre au 30 septembre)
    3. Cliquer sur **Impression**, puis **Édition CSV**
    4. Déposer ce fichier dans la zone ci-dessus
    """)
