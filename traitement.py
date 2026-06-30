"""
Logique de transformation de l'extraction VIGIEsip (.csv) vers le fichier
Excel final destiné aux agents.

Ce module ne contient que les fonctions de traitement (pas d'interface) :
il est utilisé à la fois par le script en ligne de commande
(generer_fichier_croisieres.py) et par l'application Streamlit (app.py).
"""

import datetime as dt
import io
import sys

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ---------------------------------------------------------------------------
# Correspondance entre les libellés de poste VIGIEsip et les codes courts
# utilisés dans le fichier final (cf. mode opératoire d'Yvon).
# ---------------------------------------------------------------------------
POSTE_VIGIE_VERS_CODE = {
    "POINTE SIMON EST": "CROE",
    "POINTE SIMON OUEST": "CROW",
    "TOURELLES": "TOUR",
    "QUAI DES ANNEXES": "ANNX",
}

# Postes "standards" mentionnés dans le fichier d'Yvon (GQ, SP, BEACH...)
# mais absents de l'extraction VIGIEsip : leur origine reste à confirmer
# avec Yvon/Lauriane. Dans cette version simplifiée, seules les escales
# réelles sont affichées, donc ce point n'a plus d'impact direct — mais il
# reste pertinent si un jour ces postes doivent être intégrés.

MOIS_FR = ["janvier", "février", "mars", "avril", "mai", "juin", "juillet",
           "août", "septembre", "octobre", "novembre", "décembre"]


class FichierVigieInvalide(Exception):
    """Levée quand le csv déposé ne ressemble pas à une extraction VIGIEsip valide."""


# ---------------------------------------------------------------------------
# Étape 1 : lecture et nettoyage de l'extraction VIGIEsip
# ---------------------------------------------------------------------------
def lire_extraction_vigie(fichier) -> pd.DataFrame:
    """
    Lit le csv VIGIEsip brut (encodage Windows, séparateur ';').
    `fichier` peut être un chemin (str) ou un objet fichier en mémoire
    (ex: ce que fournit st.file_uploader).
    """
    try:
        df = pd.read_csv(fichier, sep=";", encoding="cp1252")
    except UnicodeDecodeError as e:
        raise FichierVigieInvalide(
            "Le fichier ne semble pas encodé comme une extraction VIGIEsip "
            "habituelle (encodage inattendu)."
        ) from e

    df.columns = [c.strip() for c in df.columns]

    colonnes_attendues = {"N°", "Sens", "Navire", "Armateur", "Agent",
                           "Arrivée Rade", "Poste", "Dép quai", "Validé Agent"}
    manquantes = colonnes_attendues - set(df.columns)
    if manquantes:
        raise FichierVigieInvalide(
            f"Colonnes manquantes dans le fichier déposé : {sorted(manquantes)}. "
            "Ce n'est peut-être pas une extraction VIGIEsip 'Prévisions Paquebots'."
        )

    return df


def construire_escales(df: pd.DataFrame) -> pd.DataFrame:
    """
    Reconstitue une ligne par escale (arrivée + départ) à partir des paires
    de lignes DPQ / DS consécutives du CSV VIGIEsip.

    Le mode opératoire d'Yvon fait ce même travail "à la main" en décalant
    la colonne Dép quai d'une ligne vers le haut ; on reproduit la même
    logique par paires consécutives plutôt que par le N° (qui n'est pas un
    identifiant unique d'escale : un même navire peut faire plusieurs
    rotations dans la saison sous le même N°).
    """
    df = df.reset_index(drop=True)

    if len(df) % 2 != 0:
        raise FichierVigieInvalide(
            "Nombre de lignes impair dans l'extraction : impossible de "
            "former des paires DPQ/DS. Vérifier le fichier source."
        )

    arrivees = df.iloc[0::2].reset_index(drop=True)
    departs = df.iloc[1::2].reset_index(drop=True)

    sens_ok = (arrivees["Sens"] == "DPQ").all() and (departs["Sens"] == "DS").all()
    navire_ok = (arrivees["Navire"].values == departs["Navire"].values).all()
    if not (sens_ok and navire_ok):
        raise FichierVigieInvalide(
            "Les lignes du CSV ne s'enchaînent pas en paires DPQ/DS "
            "cohérentes. Le fichier source a peut-être changé de structure."
        )

    escales = pd.DataFrame({
        "Navire": arrivees["Navire"],
        "Armateur": arrivees["Armateur"],
        "Agent": arrivees["Agent"],
        "Loa": arrivees["L"],
        "Lar": arrivees["l"],
        "Te": arrivees["TE"],
        "Pax": arrivees["Pax"],
        "Crew": arrivees["Crew"],
        "ArriveeRade": pd.to_datetime(
            arrivees["Arrivée Rade"], format="%d/%m/%Y %H:%M", errors="coerce"
        ),
        "Poste": arrivees["Poste"],
        "DepQuai": pd.to_datetime(
            departs["Dép quai"], format="%d/%m/%Y %H:%M", errors="coerce"
        ),
        "ValideAgent": arrivees["Validé Agent"],
    })

    return escales


def filtrer_escales_validees(escales: pd.DataFrame) -> pd.DataFrame:
    """
    Exclut les escales dont l'agent a indiqué 'Non' (cf. mode opératoire :
    "Supprimer toutes les lignes contenant NON, il ne reste que les
    escales validées"). Les escales 'Oui' et 'Inconnu' sont conservées.
    """
    return escales[escales["ValideAgent"] != "Non"].reset_index(drop=True)


def recoder_postes(escales: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    """
    Remplace les libellés VIGIEsip par les codes courts (CROE, CROW...).
    Retourne aussi la liste des libellés non reconnus (pour avertir l'utilisateur).
    """
    escales = escales.copy()
    escales["PosteCode"] = escales["Poste"].map(POSTE_VIGIE_VERS_CODE)

    inconnus = sorted(escales[escales["PosteCode"].isna()]["Poste"].dropna().unique().tolist())
    if inconnus:
        escales.loc[escales["PosteCode"].isna(), "PosteCode"] = escales["Poste"]

    return escales, inconnus


# ---------------------------------------------------------------------------
# Étape 2 : construction du tableau final (une ligne par escale réelle)
# ---------------------------------------------------------------------------
def construire_tableau(escales: pd.DataFrame, date_debut: dt.date, date_fin: dt.date) -> pd.DataFrame:
    """
    Construit le tableau final : une ligne par escale réelle (pas de lignes
    vides pour les postes/jours sans navire), trié par date d'arrivée, avec
    une colonne "Mois" pour pouvoir filtrer facilement dans Excel.
    """
    escales = escales.copy()
    escales["JourArrivee"] = escales["ArriveeRade"].dt.date
    escales = escales[
        (escales["JourArrivee"] >= date_debut) & (escales["JourArrivee"] <= date_fin)
    ]
    escales = escales.sort_values("ArriveeRade").reset_index(drop=True)

    lignes = []
    for _, esc in escales.iterrows():
        mois_idx = esc["ArriveeRade"].month - 1
        lignes.append({
            "Date": esc["JourArrivee"],
            "Mois": MOIS_FR[mois_idx].capitalize(),
            "Poste": esc["PosteCode"],
            "Navire": esc["Navire"],
            "Compagnie": esc["Armateur"],
            "ETA": esc["ArriveeRade"],
            "ETD": esc["DepQuai"],
            "Agent": esc["Agent"],
            "Loa": esc["Loa"],
            "Lar": esc["Lar"],
            "Te": esc["Te"],
            "Pax": esc["Pax"],
            "Crew": esc["Crew"],
        })

    return pd.DataFrame(lignes)


# ---------------------------------------------------------------------------
# Étape 3 : écriture du fichier Excel mis en forme
# ---------------------------------------------------------------------------
def formater_heure(valeur):
    if pd.isna(valeur):
        return None
    return valeur.strftime("%Hh%M")


def ecrire_excel(tableau: pd.DataFrame, titre: str, date_maj: dt.date) -> bytes:
    """Construit le fichier Excel final et le retourne sous forme de bytes
    (prêt à être proposé au téléchargement par Streamlit)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Escales"

    font_titre = Font(name="Calibri", size=14, bold=True)
    font_souscription = Font(name="Calibri", size=10, italic=True, color="666666")
    font_entete = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_normal = Font(name="Calibri", size=10)
    fill_entete = PatternFill("solid", fgColor="1F4E78")
    centre = Alignment(horizontal="center")

    ws["A1"] = titre
    ws["A1"].font = font_titre

    ws["A2"] = f"À jour au {date_maj.strftime('%d/%m/%Y')} — {len(tableau)} escales"
    ws["A2"].font = font_souscription

    entetes = ["Date", "Mois", "Poste", "Navire", "Compagnie", "ETA", "ETD",
               "Agent", "Loa (m)", "Lar (m)", "Te (m)", "Pax", "Crew"]
    ligne_entete = 4
    for col_idx, libelle in enumerate(entetes, start=1):
        c = ws.cell(row=ligne_entete, column=col_idx, value=libelle)
        c.font = font_entete
        c.fill = fill_entete
        c.alignment = centre

    for i, row in enumerate(tableau.itertuples(index=False), start=ligne_entete + 1):
        eta_val = formater_heure(row.ETA) if not isinstance(row.ETA, str) else row.ETA
        etd_val = formater_heure(row.ETD) if not isinstance(row.ETD, str) else row.ETD
        valeurs = [
            row.Date, row.Mois, row.Poste, row.Navire, row.Compagnie,
            eta_val, etd_val, row.Agent, row.Loa, row.Lar, row.Te, row.Pax, row.Crew,
        ]
        for col_idx, val in enumerate(valeurs, start=1):
            c = ws.cell(row=i, column=col_idx, value=val)
            c.font = font_normal
            if col_idx == 1 and val is not None:
                c.number_format = "dd/mm/yyyy"

    derniere_ligne = ligne_entete + len(tableau)
    derniere_col = get_column_letter(len(entetes))
    plage = f"A{ligne_entete}:{derniere_col}{derniere_ligne}"

    table = Table(displayName="EscalesPaquebots", ref=plage)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(table)

    largeurs = [12, 11, 9, 25, 25, 8, 8, 10, 9, 9, 8, 8, 8]
    for col_idx, largeur in enumerate(largeurs, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = largeur

    ws.freeze_panes = f"A{ligne_entete + 1}"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# Fonction "tout-en-un" utilisée par l'appli Streamlit
# ---------------------------------------------------------------------------
def transformer(fichier_csv, date_debut: dt.date, date_fin: dt.date, titre: str):
    """
    Prend le fichier csv déposé (chemin ou objet en mémoire) et les bornes
    de saison, retourne (bytes_du_xlsx, nombre_escales, postes_non_reconnus).
    Lève FichierVigieInvalide si le fichier n'a pas le format attendu.
    """
    df_brut = lire_extraction_vigie(fichier_csv)
    escales = construire_escales(df_brut)
    escales = filtrer_escales_validees(escales)
    escales, postes_inconnus = recoder_postes(escales)

    tableau = construire_tableau(escales, date_debut, date_fin)
    contenu_xlsx = ecrire_excel(tableau, titre, date_maj=dt.date.today())

    return contenu_xlsx, len(tableau), postes_inconnus
